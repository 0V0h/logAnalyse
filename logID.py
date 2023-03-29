from xml.dom.minidom import parse
import openpyxl
from openpyxl.workbook import Workbook
import re



def write_mutype(ws,i,time,computername,domain,user,eventid,src_IP,src_port):
    ws['A{}'.format(i)] = str(time).strip()
    ws['B{}'.format(i)] = str(eventid).strip()
    ws['C{}'.format(i)] = str(src_IP).strip()
    ws['D{}'.format(i)] = str(src_port).strip()
    ws['E{}'.format(i)] = str(user).strip()
    ws['F{}'.format(i)] = str(computername).strip()
    ws['G{}'.format(i)] = str(domain).strip()



def id_4624(file="./result/4624.xml"):
    # 读取文件
    dom = parse(file)
    # 获取文档元素对象
    data = dom.documentElement
    # 获取 student
    rows = data.getElementsByTagName('row')

    workbook = Workbook()
    workbook.save('./result/4624.xlsx')
    wb = openpyxl.load_workbook("./result/4624.xlsx")
    Sheet = wb['Sheet']
    wb.remove(Sheet)
    smb = wb.create_sheet('smb登录')
    rdp = wb.create_sheet('RDP登录')
    console = wb.create_sheet('控制台登录')
    other = wb.create_sheet('其他登录')
    # ws = wb["Sheet"]
    sheet = []
    sheet.append(smb)
    sheet.append(rdp)
    sheet.append(console)
    sheet.append(other)
    for ws in sheet:
        ws['A1'] = '登录时间'
        ws['B1'] = '日志类型'
        ws['C1'] = '登录源IP'
        ws['D1'] = '登录源端口'
        ws['E1'] = '登录账户'
        ws['F1'] = '主机名'
        ws['G1'] = 'domain'


    smb = 2
    rdp = 2
    console = 2
    other = 2


    for row in rows:
        
        time = row.getElementsByTagName('timegenerated')[0].childNodes[0].nodeValue
        computername = row.getElementsByTagName('computername')[0].childNodes[0].nodeValue
        strings = row.getElementsByTagName('strings')[0].childNodes[0].nodeValue
        
        strings = str(strings).strip()
        try:
            reg = re.compile("(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|.*")
            group = reg.findall(strings)[0]
        except:
            reg = re.compile("(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*?)\|(.*)")
            group = reg.findall(strings)[0]

        domain = group[2]
        user = group[5]
        eventid = group[8]
        src_IP = group[18]
        src_port = group[19]


        if eventid == "3":
            write_mutype(sheet[0],smb,time,computername,domain,user,eventid,src_IP,src_port)
            smb += 1
        elif eventid == "10":
            write_mutype(sheet[1],rdp,time,computername,domain,user,eventid,src_IP,src_port)
            rdp += 1
        elif eventid == "2":
            write_mutype(sheet[2],console,time,computername,domain,user,eventid,src_IP,src_port)
            console += 1
        else :
            write_mutype(sheet[3],other,time,computername,domain,user,eventid,src_IP,src_port)
            other += 1
        

    wb.save("./result/4624.xlsx")
    wb.close()







def id_4625(file="./result/4625.xml"):
    # 读取文件
    dom = parse(file)
    # 获取文档元素对象
    data = dom.documentElement
    # 获取 student
    rows = data.getElementsByTagName('row')

    workbook = Workbook()
    workbook.save('./result/4625.xlsx')
    wb = openpyxl.load_workbook("./result/4625.xlsx")
    Sheet = wb['Sheet']
    wb.remove(Sheet)
    ws = wb.create_sheet('登录失败')

    # ws = wb["Sheet"]

    ws['A1'] = '登录时间'
    ws['B1'] = '用户名'
    ws['C1'] = '主机名'
    ws['D1'] = '源IP'
    ws['E1'] = '源端口'
    ws['F1'] = '报错信息'


    err = {
        "0xc0000064":"用户名不存在",
        "0xc000006a":"用户名是正确的,但密码是错误的",
        "0xc0000234":"用户当前锁定",
        "0xc0000072":"帐户目前禁用",
        "0xc000006f":"用户试图登录天的外周或时间限制",
        "0xc0000070":"工作站的限制",
        "0xc0000193":"帐号过期",
        "0xc0000071":"过期的密码",
        "0xc0000133":"时钟之间的直流和其他电脑太不同步",
        "0xc0000224":"在下次登录用户需要更改密码",
        "0xc0000225":"显然一个缺陷在Windows和不是一个风险",
        "0xc000015b":"没有被授予该用户请求登录类型(又名登录正确的)在这台机器",
        "0xc000006d":"似乎是由于系统问题和不安全"
    }


    i = 1
    for row in rows:
        
        i += 1

        time = row.getElementsByTagName('timegenerated')[0].childNodes[0].nodeValue
        computername = row.getElementsByTagName('computername')[0].childNodes[0].nodeValue
        strings = row.getElementsByTagName('strings')[0].childNodes[0].nodeValue
    
        strings = str(strings).strip()
        reg = re.compile(".*?\|.*?\|.*?\|.*?\|.*?\|(.*?)\|.*?\|.*?\|.*?\|(.*?)\|.*?\|.*\|.*\|.*\|.*\|.*\|.*\|.*\|.*\|(.*?)\|(.*)")
        group = reg.findall(strings)[0]

        user = group[0]
        err_msg = group[1]
        src_IP = group[2]
        src_port = group[3]

        ws['A{}'.format(i)] = str(time).strip()
        ws['B{}'.format(i)] = str(user).strip()
        ws['C{}'.format(i)] = str(computername).strip()
        ws['D{}'.format(i)] = str(src_IP).strip()
        ws['E{}'.format(i)] = str(src_port).strip()
        ws['F{}'.format(i)] = f"{str(err_msg).strip()}({err[str(err_msg).strip()]})"

    wb.save("./result/4625.xlsx")
    wb.close()
        







def id_1102(file="./result/1102.xml"):
    # 读取文件
    dom = parse(file)
    # 获取文档元素对象
    data = dom.documentElement
    # 获取 student
    rows = data.getElementsByTagName('row')

    workbook = Workbook()
    workbook.save('./result/1102.xlsx')
    wb = openpyxl.load_workbook("./result/1102.xlsx")
    Sheet = wb['Sheet']
    wb.remove(Sheet)
    ws = wb.create_sheet('日志清除')

    # ws = wb["Sheet"]

    ws['A1'] = '登录时间'
    ws['B1'] = '操作用户'
    ws['C1'] = '主机名'
    ws['D1'] = '登录类型'


    i = 1
    for row in rows:
        
        i += 1

        time = row.getElementsByTagName('timegenerated')[0].childNodes[0].nodeValue
        computername = row.getElementsByTagName('computername')[0].childNodes[0].nodeValue
        strings = row.getElementsByTagName('strings')[0].childNodes[0].nodeValue
    
        strings = str(strings).strip()


        reg = re.compile(".*?\|(.*?)\|(.*?)\|(.*)")
        group = reg.findall(strings)[0]

        user = group[0]
        logtype = group[2]


        ws['A{}'.format(i)] = str(time).strip()
        ws['B{}'.format(i)] = str(user).strip()
        ws['C{}'.format(i)] = str(computername).strip()
        ws['D{}'.format(i)] = str(logtype).strip()

    wb.save("./result/1102.xlsx")
    wb.close()





def id_1149(file="./result/1149.xml"):
    # 读取文件
    dom = parse(file)
    # 获取文档元素对象
    data = dom.documentElement
    # 获取 student
    rows = data.getElementsByTagName('row')

    workbook = Workbook()
    workbook.save('./result/1149.xlsx')
    wb = openpyxl.load_workbook("./result/1149.xlsx")
    Sheet = wb['Sheet']
    wb.remove(Sheet)
    ws = wb.create_sheet('RDP应用登录')

    # ws = wb["Sheet"]

    ws['A1'] = '登录时间'
    ws['B1'] = '登录用户'
    ws['C1'] = '主机名'
    ws['D1'] = '登录源IP'


    i = 1
    for row in rows:
        
        i += 1

        time = row.getElementsByTagName('timegenerated')[0].childNodes[0].nodeValue
        computername = row.getElementsByTagName('computername')[0].childNodes[0].nodeValue
        strings = row.getElementsByTagName('strings')[0].childNodes[0].nodeValue
    
        strings = str(strings).strip()
        reg = re.compile("(.*?)\|(.*?)\|(.*)")
        group = reg.findall(strings)[0]


        user = group[0]
        src_IP = group[2]


        ws['A{}'.format(i)] = str(time).strip()
        ws['B{}'.format(i)] = str(user).strip()
        ws['C{}'.format(i)] = str(computername).strip()
        ws['D{}'.format(i)] = str(src_IP).strip()

    wb.save("./result/1149.xlsx")
    wb.close()








def id_6005(file="./result/6005.xml"):
    # 读取文件
    dom = parse(file)
    # 获取文档元素对象
    data = dom.documentElement
    # 获取 student
    rows = data.getElementsByTagName('row')

    workbook = Workbook()
    workbook.save('./result/6005.xlsx')
    wb = openpyxl.load_workbook("./result/6005.xlsx")
    Sheet = wb['Sheet']
    wb.remove(Sheet)
    ws = wb.create_sheet('日志启动')

    # ws = wb["Sheet"]

    ws['A1'] = '登录时间'
    ws['B1'] = 'event信息'
    ws['C1'] = '主机名'


    i = 1
    for row in rows:
        
        i += 1

        time = row.getElementsByTagName('timegenerated')[0].childNodes[0].nodeValue
        computername = row.getElementsByTagName('computername')[0].childNodes[0].nodeValue
        message = row.getElementsByTagName('message')[0].childNodes[0].nodeValue
    

        ws['A{}'.format(i)] = str(time).strip()
        ws['B{}'.format(i)] = str(message).strip()
        ws['C{}'.format(i)] = str(computername).strip()

    wb.save("./result/6005.xlsx")
    wb.close()





def id_25(file="./result/25.xml"):
    # 读取文件
    dom = parse(file)
    # 获取文档元素对象
    data = dom.documentElement
    # 获取 student
    rows = data.getElementsByTagName('row')

    workbook = Workbook()
    workbook.save('./result/25.xlsx')
    wb = openpyxl.load_workbook("./result/25.xlsx")
    Sheet = wb['Sheet']
    wb.remove(Sheet)
    ws = wb.create_sheet('RDP重连')

    # ws = wb["Sheet"]

    ws['A1'] = '登录时间'
    ws['B1'] = '用户'
    ws['C1'] = '类型'
    ws['D1'] = '源IP'


    i = 1
    for row in rows:
        
        i += 1

        time = row.getElementsByTagName('timegenerated')[0].childNodes[0].nodeValue
        computername = row.getElementsByTagName('computername')[0].childNodes[0].nodeValue
        strings = row.getElementsByTagName('strings')[0].childNodes[0].nodeValue
    
        strings = str(strings).strip()
        reg = re.compile("(.*?)\|(.*?)\|(.*)")
        group = reg.findall(strings)[0]

        user = group[0]
        type = group[1]
        src_IP = group[2]

        ws['A{}'.format(i)] = str(time).strip()
        ws['B{}'.format(i)] = str(user).strip()
        ws['C{}'.format(i)] = str(type).strip()
        ws['D{}'.format(i)] = str(src_IP).strip()

    wb.save("./result/25.xlsx")
    wb.close()






def id_104(file="./result/104.xml"):
    # 读取文件
    dom = parse(file)
    # 获取文档元素对象
    data = dom.documentElement
    # 获取 student
    rows = data.getElementsByTagName('row')

    workbook = Workbook()
    workbook.save('./result/104.xlsx')
    wb = openpyxl.load_workbook("./result/104.xlsx")
    Sheet = wb['Sheet']
    wb.remove(Sheet)
    ws = wb.create_sheet('日志清除')

    # ws = wb["Sheet"]

    ws['A1'] = '登录时间'
    ws['B1'] = 'event信息'
    ws['C1'] = '主机名'
    ws['D1'] = '操作用户'


    i = 1
    for row in rows:
        
        i += 1

        time = row.getElementsByTagName('timegenerated')[0].childNodes[0].nodeValue
        computername = row.getElementsByTagName('computername')[0].childNodes[0].nodeValue
        message = row.getElementsByTagName('message')[0].childNodes[0].nodeValue
        strings = row.getElementsByTagName('strings')[0].childNodes[0].nodeValue
    
        strings = str(strings).strip()
        reg = re.compile("(.*?)\|(.*?)\|(.*)")
        group = reg.findall(strings)[0]

        user = group[0]
    
        ws['A{}'.format(i)] = str(time).strip()
        ws['B{}'.format(i)] = str(message).strip()
        ws['C{}'.format(i)] = str(computername).strip()
        ws['D{}'.format(i)] = str(user).strip()

    wb.save("./result/104.xlsx")
    wb.close()










def id_4720(file="./result/4720.xml"):
    # 读取文件
    dom = parse(file)
    # 获取文档元素对象
    data = dom.documentElement
    # 获取 student
    rows = data.getElementsByTagName('row')

    workbook = Workbook()
    workbook.save('./result/4720.xlsx')
    wb = openpyxl.load_workbook("./result/4720.xlsx")
    Sheet = wb['Sheet']
    wb.remove(Sheet)
    ws = wb.create_sheet('创建用户')

    # ws = wb["Sheet"]

    ws['A1'] = '登录时间'
    ws['B1'] = '创建用户'
    ws['C1'] = '主机名'
    ws['D1'] = '操作用户'


    i = 1
    for row in rows:
        
        i += 1

        time = row.getElementsByTagName('timegenerated')[0].childNodes[0].nodeValue
        strings = row.getElementsByTagName('strings')[0].childNodes[0].nodeValue
    
        strings = str(strings).strip()
        reg = re.compile("(.*?)\|.*?\|.*?\|.*?\|(.*?)\|(.*?)\|.*?\|.*?\|.*?\|.*?\|.*?\|.*?\|.*?\|.*?\|.*?\|.*?\|.*?\|.*?\|.*?\|.*?\|.*?\|.*?\|")
        group = reg.findall(strings)[0]

        add_usr = group[0]
        operator = group[1]
        computername = group[2]
    
        ws['A{}'.format(i)] = str(time).strip()
        ws['B{}'.format(i)] = str(add_usr).strip()
        ws['C{}'.format(i)] = str(computername).strip()
        ws['D{}'.format(i)] = str(operator).strip()

    wb.save("./result/4720.xlsx")
    wb.close()





